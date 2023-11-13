import os
import sys
from datetime import datetime, timedelta as td

import pytz
import openpyxl
import pandas as pd
import numpy
from openpyxl import load_workbook

user_dir = os.environ.get('USERPROFILE')
user_docs = os.path.join(user_dir,"Documents")
code_dir = os.path.join(user_docs,"code")
pyexamples_dir = os.path.join(code_dir,"Python examples")
utils_dir = os.path.join(pyexamples_dir,"Utils")
sys.path.append(utils_dir)
#  My User LIbrary setup ends
import where_is
# import config

# def find_all(name, path):
#     result = []
#     for root, dirs, files in os.walk(path):
#         if name in files:
#             result.append(os.path.join(root, name))
#     return result

def iterating_over_values(path, sheet_name, cell_range):
    workbook = load_workbook(filename=path)
    if sheet_name not in workbook.sheetnames:
        print(f"'{sheet_name}' not found. Quitting.")
        return
    sheet = workbook[sheet_name]
    for column in sheet[cell_range]:
        for cell in column:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                # Skip this cell
                continue
            print(f"{cell.column_letter}{cell.row} = {cell.value}")

def read_flights(file_name, bookings, excel_df):

    excel_df = pd.read_excel(file_name, sheet_name=bookings)
    return (excel_df)

def df_functions_panda(df):
    print("Data Frame stuff")
    print ("Start of df functions")
    print ("Count :", df.count())
    print ("\nHead :\n", df.head(3))
    print ("\nDescribe :\n", df.describe())
    labels = df.columns.values
    print ("\nColumn Labels\n")
    print (labels)
    print("\nData Frame stuff END \n")

def stripTimeDate(input_str, m, d, h, mm, ampm):
    mth = input_str.split(",")
    m, d = mth[1].split()

    hrmmAMPM = input_str.split(",")[0]
    h = hrmmAMPM.split(":")[0]
    mm_tmp = hrmmAMPM.split(":")[1]
    mm = mm_tmp[:2].strip()
    # print ("mmTmp :", mm_tmp)
    ampm = mm_tmp[2:4].strip()
    # print(ampm)
    return (m,d,h,mm,ampm)

def convertDateStr(d):
    mthno = {
        'Jan': '01',
        'Feb': '02',
        'Mar': '03',
        'Apr': '04',
        'May': '05',
        'Jun': '06',
        'Jul': '07',
        'Aug': '08',
        'Sep': '09',
        'Oct': '10',
        'Nov': '11',
        'Dec': '12'
    }
    return (mthno[d])

def stepThrough_df(df):
    formatter_string = "%d-%m-%y %h:%m"

    dstr = ""
    yr = datetime.now().year

    rowCount = df.shape[0]
    colCount = df.shape[1]

    # Stuff
    # print ("Shape 1 = ", colCount)
    # print (df.iloc[7,1])

    # for c in df.columns:
    #     print (c," / ", end = "")
    #
    # for r in range(0,rowCount):
    #     for c in  range(0, colCount):
    #         print (df.iloc[r,c], end = '')
    #
    # print ("\n")

    for r in range(0,rowCount):
        for c in  range(5, 7):

            input_str, mth_of, day_of, hr, mm, ampm = ["","","","","",""]

            if c == 5:
                # print ("Departure Time :{}  {}".format(df.iloc[r,c].strip(), df.iloc[r,c+2]))
                input_str =  df.iloc[r, c]
                mth_of, day_of, hr, mm, ampm = stripTimeDate(input_str,mth_of,day_of, hr, mm, ampm)

                mth_of =  convertDateStr(mth_of)
                hr = hr if ampm == "AM" else str(int(hr) + 12)
                dstr = day_of + "/" + mth_of + "/" + str(yr) + " " + hr + ":" + mm
                if hr == "24":
                    hr,mm = "23", "59"

                # print (dstr)
                # print (yr, int(mth_of), int(day_of), int(hr), int(mm) )
                datetime_object = datetime(yr, int(mth_of), int(day_of), int(hr), int(mm))
                if hr == "23" and mm == "59":
                    # print ("***",datetime_object)
                    t_obj = td(minutes=1)
                    datetime_object = datetime_object + t_obj
                    # print ("***",datetime_object)
                print(datetime_object)
                # # print("Month <", mth_of, ">\nDay of <", day_of, ">\nHours <{}>\nmin <{}>".format(hr,mm), end="\n")

            if c == 6:
                # print ("Arrival Time :{}  {}".format(df.iloc[r,c].strip(), df.iloc[r,c+2]))
                input_str =  df.iloc[r, c]
                mth_of, day_of, hr, mm, ampm = stripTimeDate(input_str,mth_of,day_of, hr, mm, ampm)

                mth_of =  convertDateStr(mth_of)
                hr = hr if ampm == "AM" else str(int(hr) + 12)
                dstr = day_of + "/" + mth_of + "/" + str(yr) + " " + hr + ":" + mm
                if hr == "24":
                    hr, mm = "23", "59"

                print(dstr)
                # datetime_object = datetime(yr,int(mth_of),int(day_of),int(hr),int(mm))

                if hr == "23" and mm == "59":
                    # print("^^^", datetime_object)
                    t_obj = td(minutes=1)
                    datetime_object = datetime_object + t_obj
                    # print("^^^", datetime_object)
                print (datetime_object)
            # print ("r = {} {}".format(r, df.iloc[r,1]))
        print ("\n")

def tz_stuff():
    # current Datetime
    unaware = datetime.now()
    print('Timezone naive:', unaware)

    # Standard UTC timezone aware Datetime
    aware = datetime.now(pytz.utc)
    print('Timezone Aware:', aware)

    # US/Central timezone datetime
    aware_us_central = datetime.now(pytz.timezone('US/Central'))
    print('US Central DateTime', aware_us_central)

    print (pytz.common_timezones)
    for i in pytz.common_timezones:
        print (i)



if __name__ == "__main__":
    xl_file_loc = where_is.file_path('Flights.xlsx')
    # print(where_is.file_path('Flights.xlsx'))
    # print("*****************************)")

    excel_df = pd.DataFrame


    world_tzs = (
    'America/Los_Angeles', 'America/Toronto', 'Asia/Singapore', 'Asia/Taipei', 'Australia/Brisbane', 'Australia/Sydney','Europe/London')
    # world_tzs = {Location = 'Los Angeles': Value = 'America/Los_Angeles'}


    for i in world_tzs:
        unaware = datetime.now()
        othertime = datetime.now(pytz.timezone(i))
        # print(i, othertime)

    # iterating_over_values(xl_file_loc, sheet_name="Bookings", cell_range="A2:H9")
    excel_df = read_flights(xl_file_loc,"Bookings",excel_df)

    df_functions_panda(excel_df)

    print (excel_df)

    stepThrough_df(excel_df)

    print(convertDateStr("Apr"))
    print(convertDateStr("Jan"))
    print(convertDateStr("Nov"))
    # for i in excel_df.columns.
    #
    # for i in world_tzs:
    #     print (i)
    #
    # # for column in excel_df:
    # #     # print(type(column_name))
    # #     print(column)
    # #     # print('------\n')
    #
    # excel_df['Australia dt'] = ''
    # excel_df['Australia at'] = ''
    # excel_df['United Kingdom dt'] = ''
    # excel_df['United Kingdom at'] = ''
    #
    # for column_names in excel_df:
    #     print (column_names)
    #
    # for dt, tz in zip(excel_df['Departure time'], excel_df['Departure TZ']):
    #     print(dt, tz)
    #     if tz == '':
    #         print (tz)
    #
    # # print (excel_df.columns)
    # for c in excel_df.columns:
    #     print ("ask > ", c)
    #
    # print ('Shape')
    # print ("Data frame has ", excel_df.shape[0], " rows", end = "")
    # print (" and ", excel_df.shape[1], " columns", end="\n")
    # print (excel_df.columns.set_names, "\n")
    # # for row in excel_df:
    # #     print ("---", row)
    #
    # for r in range(0,excel_df.shape[0]):
    #     ss = excel_df.iloc[[r]]['Departure time']
    #     # print (excel_df.iloc[[r]]['Departure time'])
    #     print (type(ss))
    #     print ("X")
    #     string_p = str(ss)
    #     print (string_p)
    #     print ("Y")
    #     s_int = str(string_p).find("AM", 0)
    #     print("Z")
    #     time_o = datetime.strptime(str(string_p), '%H:%M%p')
    #
    #     # print (time_o)
    #     print(str(string_p), time_o)
    # # ,datetime.time(excel_df['Departure time']))